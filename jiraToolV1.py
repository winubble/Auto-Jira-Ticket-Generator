#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# JIRA libraries
from jira.resources import IssueLink
import jira.client
from jira.client import JIRA
from jira import JIRA
import re

# Excel libraries
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import openpyxl
import pandas as pd

# GUI libraries
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter.filedialog import askopenfilename
import tkinter.scrolledtext as tkst
from tkinter import messagebox as mb
from tkinter import filedialog 
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""Variables"""
username = 'Empty'
password = 'Empty'
name = 'Empty'
workbookTitle = 'Empty'
worksheetTitle = 'Empty'
sheet = 'Empty'
book = 'Empty'
__summaryTitle = 'Empty'
created_tickets = 0
updated_issues = 0
ch = 'A'

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
""" Log in window that asks for account Username and Password 
    Saves the user's username and password in a variable"""

log_inWindow = tk.Tk()
log_inWindow.title('Login JIRA Account')
log_inWindow.geometry("390x135")
# log_inWindow.resizable(False, False)
log_inWindow['bg'] = "#000000"

tk.Label(log_inWindow, text = 'Username', font = 'bold', bg = "#000000", fg = "white").grid(row = 0, column = 0, padx = 10, pady = 10)
tk.Label(log_inWindow, text = 'Password', font = 'bold', bg = "#000000", fg = "white").grid(row = 1, column = 0, padx = 10, pady = 10)

e1 = tk.Entry(log_inWindow, font = "bold", bg = "#000000", fg = "white", cursor = "heart", insertbackground = 'white')
e2 = tk.Entry(log_inWindow, show = '*', font = "bold", bg = "#000000", fg = "white", cursor = "heart", insertbackground = 'white')

e1.grid(row = 0, column = 1)
e2.grid(row = 1, column = 1)

def save_textvariable_enter(event):
    global username
    username = e1.get()
    global password
    password = e2.get()
    log_inWindow.destroy()
def save_textvariable():
    global username
    username = e1.get()
    global password
    password = e2.get()
    log_inWindow.destroy()


log_inWindow.bind('<Return>', save_textvariable_enter)
tk.Button(log_inWindow, text = 'Log In', font = "bold", command = save_textvariable, bd = 4, width = 8, bg = "#000000", fg = "white", activeforeground = "gray").grid(row = 2, column = 2, padx = 10,)

log_inWindow.mainloop()

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""Creates main GUI and asks for user to browse and select excel workbook to be used
    and Name of spreadsheet to read information from
 
    Lets user choose excel file from file explorer and only accepts openpyxl file formats"""

initWindow = tk.Tk()

initWindow.title('Open Excel File')
initWindow.geometry("500x300")
initWindow.resizable(False, False)
initWindow['bg'] = "#000000"

labelframe1 = LabelFrame(initWindow, text="Choose Excel Workbook", bg = "#000000", fg = "white")
labelframe1.pack(fill= "both", expand="yes", padx = 10, pady = 10) 

labelframe2 = LabelFrame(initWindow, text = "Select Spreadsheet", bg = "#000000", fg = "white")
labelframe2.pack(fill = "both", expand = "yes", padx = 10, pady = 10)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def OpenFile():
    """ Open File explorer and lets user select exsisting excel workbook and worksheet to be used """
    global name
    name = filedialog.askopenfilename(initialdir = "C:",
                           filetypes = (("Excel Workbook", "*.xlsx"), ("Excel Macro-Enabled Workbook","*.xlsm")),
                           title = "Choose a file."
                           )
    global workbookTitle
    workbookTitle = os.path.basename(name)
    wbkTitle = StringVar()
    wbkTitle.set(workbookTitle)

    Entry(labelframe1, textvariable = wbkTitle, state = DISABLED, width = 20, font = 'bold').place(x = 20, y = 25)

    wb = load_workbook(filename = workbookTitle)   
    n = len(wb.sheetnames)

    OPTIONS = []
    for k in range(0, n):
        OPTIONS += [wb.sheetnames[k]]

    global worksheetTitle    
    worksheetTitle = StringVar(initWindow)
    worksheetTitle.set(OPTIONS[0])

    OptionMenu(labelframe2, worksheetTitle, *OPTIONS).place(x = 20, y = 25)

def quit_initWindow_enter(event):
    initWindow.destroy()
def quit_initWindow():
    initWindow.destroy()


tk.Button(labelframe1, text = "Browse", command = OpenFile, font = 'bold', bd = 4, width = 10, bg = "#000000", fg = "white", activeforeground = "gray").place(x = 350, y = 25)
tk.Button(labelframe2, text = "Ok", command = quit_initWindow, font = 'bold', bd = 4, width = 10, bg = "#000000", fg = "white", activeforeground = "gray").place(x = 350, y = 25)
initWindow.bind('<Return>', quit_initWindow_enter)

initWindow.mainloop()
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# Save selected workbook and worksheet titles to variables
book = openpyxl.load_workbook(workbookTitle)
sheet = book.get_sheet_by_name(worksheetTitle.get())

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""Creates the main window where the selecting, creating, and updating issues are done"""
mainWindow = tk.Tk()
mainWindow.title("JIRA Import Tool")
mainWindow.state('zoomed')

mainWindow.columnconfigure(0, weight = 1)
mainWindow.rowconfigure(0, weight = 1)

canvas = tk.Canvas(mainWindow) 
frame = tk.Frame(canvas)

total_column = sheet.max_column
total_row = sheet.max_row

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def create_new_issue():
    """Creates a new issue from data entered in excel and selected using the checkboxes 
    by the user"""

    # Authentication done by using username and password
    jira = JIRA(
        basic_auth = (username, password),
        options = {'server': 'http://142.104.193.65:8080'}
        #options = {'server': 'https://jira.oceannetworks.ca/'}
    )

    items = []
    for ctr, int_var in enumerate(cb_intvar):
        if int_var.get(): 
            x = ctr+2
            # Summary 
            __instrument = sheet["I%d" % x].value           # Column I
            __serialNumber = sheet["J%d" % x].value         # Column J
            __deviceID = sheet["K%d" % x].value             # Column K

            # Descriptions
            __location = sheet["A%d" % x].value             # Column A
            __communications = sheet["D%d" % x].value       # Column D
            __cableID = sheet["F%d" % x].value              # Column F
            __instrumentCategory = sheet["H%d" % x].value   # Column H
            __ipAddress = sheet["G%d" % x].value            # Column G
            
            # Components
            __components = sheet["L%d" % x].value           # Column L           

            # Outward Link
            __outwardIssue = sheet["N%d" % x].value         # Column N

            __summaryTitle = sheet["O%d" % x].value         # Column O   

            if __summaryTitle == "Recovery":
                __summaryTitle = "Instrument Receiving/Bench Test"
            elif __summaryTitle == "Deploy":
                __summaryTitle = "Instrument Qualification"

            print(__deviceID)
            # create new jira ticket
            #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
            new_issue = jira.create_issue(
                project = {'key': 'EN'}, 
                summary = "%s: %s SI: %s DI: %s" % (__summaryTitle, __instrument, __serialNumber, __deviceID),
                description = "Site Location: %s\nCommunications: %s\nCable EXT ID: %s\n%s\n%s" % (__location, __communications, __cableID, __instrumentCategory, __ipAddress), 
                issuetype = {'name': 'Task'}, 
                components = [{'name' : __components}],
                customfield_10794 = {'id': "10453"},            # Bill of work to Customers (Default: ONC Internal)
                customfield_10592 = "%s" % __serialNumber,      # Serial # field
                customfield_10070 = __deviceID)                 # Device ID field

            # If Issue Link column is filled then create issue link
            if __outwardIssue != None:
                """create issue linkfor the newly created issue"""
                # create_issue_link(type, inwardIssue, outwardIssue, comment)
                jira.create_issue_link("Related", new_issue.key, __outwardIssue, None)

            # Displays the number of Issues created
            items.append(new_issue)
            global created_tickets
            created_tickets  = len(items)

            # Print new issue key created to designated cell in excel spredsheet
            #sheet["L%d" % x] = "https://jira.oceannetworks.ca/browse/%s" % new_issue.key
            
            # Print newly created issue key as a hyperlink
            sheet["M%d" % x] = "http://142.104.193.65:8080/browse/%s" % new_issue.key

            # Excel spreadsheet must be closed in order to save file
            book.save(workbookTitle)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def update_issue():
    """Updates issue summary, description, and issue links (is there are any)
    selected by the user using the checkboxes"""

    # Authentication done by using username and password
    jira = JIRA(
        basic_auth = (username, password),
        options = {'server': 'http://142.104.193.65:8080'}
        #options = {'server': 'https://jira.oceannetworks.ca/'}
    )
    items = []
    for ctr, int_var in enumerate(cb_intvar):
        if int_var.get(): 
            x = ctr+2
            # Summary 
            __instrument = sheet["I%d" % x].value           # Column I
            __serialNumber = sheet["J%d" % x].value         # Column J
            __deviceID = sheet["K%d" % x].value             # Column K

            # Descriptions
            __location = sheet["A%d" % x].value             # Column A
            __communications = sheet["D%d" % x].value       # Column D
            __cableID = sheet["F%d" % x].value              # Column F
            __instrumentCategory = sheet["H%d" % x].value   # Column H
            __ipAddress = sheet["G%d" % x].value            # Column G
            
            # Components
            __components = sheet["L%d" % x].value           # Column L

            # Issue Key
            __issueKey = sheet["M%d" % x].value             # Column M

            # Outward Issue Link
            __outwardIssue = sheet["N%d" % x].value         # Column N

            # To extract the issue key from the link
            # EN-XXXX = [-7:]
            # EN-XXXXX = [-8:]
            findEN = 'EN'
            #__inwardIssue = __issueKey[-8:]
            __inwardIssue = __issueKey[__issueKey.find(findEN) : ]

            # Issue to be updated
            issue = jira.issue(__inwardIssue)

            __summaryTitle = sheet["O%d" % x].value         # Column O

            if __summaryTitle == "Recovery":
                __summaryTitle = "Instrument Receiving/Bench Test"
            elif __summaryTitle == "Deploy":
                __summaryTitle = "Instrument Qualification"


            # # update selected issue
            issue.update(
                project = {'key': 'EN'}, 
                summary = "%s: %s %s %s" % (__summaryTitle, __instrument, __serialNumber, __deviceID),
                description = "Site Location: %s\nCommunications: %s\nCable EXT ID: %s\n%s\n%s" % (__location, __communications, __cableID, __instrumentCategory, __ipAddress), 
                customfield_10794 = {'id': "10453"},            # Bill of worl to Customers (Default: ONC Internal)
                customfield_10592 = "%s" % __serialNumber,      # Serial # field (Should be string)
                customfield_10070 = __deviceID)                 # Device ID field

            # if Issue Link column is filled then create issue link
            if __outwardIssue != None:
                jira.create_issue_link("Related", issue, __outwardIssue, None)

            # to display the number of issues updated
            items.append(issue)
            global updated_issues 
            updated_issues = len(items)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def test_create_new_issue():
    """Creates a dialogbox to confirm that the new issue(s) is/are created""" 
    result = mb.askquestion('Verify JIRA Issues to be Created', 'Are you sure you want to create JIRA tickets from the selected data?')
    if result == 'yes':
        create_new_issue()
        if created_tickets == 0:
            mb.showerror("Error", "Creation Unsuccessfull")
        else:
            mb.showinfo('Issue(s) Created', '%d JIRA Issue(s) were Created Successfully' % created_tickets)
    else:
        pass
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def test_update_issue():
    """Creates a dialogbox to confirm that the selected issue(s) is/are updated"""
    result = mb.askquestion('Verify JIRA Issues to be Created', 'Are you sure you want to update selected JIRA issue(s)?')
    if result == 'yes':
        update_issue()
        if updated_issues == 0:
            mb.showerror("Error", "Update Unsuccessfull")
        else:
            mb.showinfo(' Issue(s) Updated', '%d JIRA Issue(s) were Updated Successfully'  % updated_issues)
    else:
        pass
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

create_button = tk.Button(frame, text = "Create JIRA Issue",  command = test_create_new_issue, bd = 4, relief = RAISED, width = 15, height = 2)
update_button = tk.Button(frame, text = "Update JIRA Issues", command = test_update_issue, bd = 4, relief = RAISED, width = 15, height = 2)

create_button.grid(row = 0, column = 1, padx = 15, pady = 10)
update_button.grid(row = 0, column = 2, padx = 15, pady = 10)

# Excel data table
for j in range(1, total_row+1):  
    for i in range(0, total_column):
        datacolumn = chr(ord(ch) + i)
        data = "%s%d" % (datacolumn, j)
        a = sheet[data].value
        tk.Label(frame, text = a, font = 'bold').grid(row = j, column = i+1, padx = 10, pady = 10)

 # Checkboxes created to select row of data to grab data from        
cb_intvar = []  
for j in range(1, total_row):  
    cb_intvar.append(IntVar())
    chbx = tk.Checkbutton(frame, variable=cb_intvar[-1])
    chbx.grid(row = j+1, column = 0, sticky = 'w')

# select_all = tk.Button(frame, text = "Select All", command = chbx.select, bd = 4, relief = RAISED, width = 15, height = 2)
# deselect_all = tk.Button(frame, text = 'Deselect All', command = chbx.deselect, bd = 4, relief = RAISED, width = 15, height = 2)

# select_all.grid(row = 0, column = 3, padx = 15, pady = 10)
# deselect_all.grid(row = 0, column = 4, padx = 15, pady = 10)

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""Block of code for the scrollbar"""

def on_resize(event):
    """Resize canvas scrollregion when the canvas is resized."""
    canvas.configure(scrollregion=canvas.bbox('all'))

canvas.create_window(0, 0, anchor = 'nw', window = frame)

vbar = ttk.Scrollbar(mainWindow, orient = 'vertical', command = canvas.yview)
hbar = ttk.Scrollbar(mainWindow, orient = 'horizontal', command = canvas.xview)
canvas.configure(xscrollcommand = hbar.set,
                yscrollcommand = vbar.set,
                scrollregion = canvas.bbox('all'))

canvas.grid(row = 0, column = 0, sticky = 'eswn')
vbar.grid(row = 0, column = 1, sticky = 'ns')
hbar.grid(row = 1, column = 0, sticky = 'ew')

canvas.bind('<Configure>', on_resize)

mainWindow.mainloop()