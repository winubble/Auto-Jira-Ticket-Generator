"""
Author: Lu Lu
Date: 2020-12-16

This tool is used to generate jira tickets automatically
Default link to the Jira site in use

WITH UPDATE
WITH DETELE DECIMAL and RECOVERY
"""

# Excel libraries
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import openpyxl
import pandas as pd
import numpy as np

# GUI libraries
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkinter.filedialog import askopenfilename
import tkinter.scrolledtext as tkst
from tkinter import messagebox as mb
from tkinter import filedialog
import tkinter.font as tkFont

# JIRA libraries
from jira.resources import IssueLink
import jira.client
from jira.client import JIRA
from jira import JIRA
import re

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""Variables"""
username = ''
password = ''
mypath = ''
workbookTitle = ''

# create a dataframe for storing the deploy row
df_out = pd.DataFrame()

# create a dictionary for storing the site ticket EN number
site_dict = {}

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~First Window
""" Log in window that asks for account Username and Password 
    Saves the user's username and password in a variable"""

log_inWindow = tk.Tk()
log_inWindow.title('JIRA Login')
log_inWindow.geometry("390x135")
log_inWindow['bg'] = "#0E69C1"

tk.Label(log_inWindow, text = 'Username', font = 'bold', bg = "#0E69C1", fg = "white").grid(row = 0, column = 0, padx = 10, pady = 10)
tk.Label(log_inWindow, text = 'Password', font = 'bold', bg = "#0E69C1", fg = "white").grid(row = 1, column = 0, padx = 10, pady = 10)

e1 = tk.Entry(log_inWindow, font = "bold", bg = "white", fg = "black", cursor = "heart", insertbackground = 'black')
e2 = tk.Entry(log_inWindow, show = '*', font = "bold", bg = "white", fg = "black", cursor = "heart", insertbackground = 'black')

e1.grid(row = 0, column = 1)
e2.grid(row = 1, column = 1)

def save_textvariable():
    global username
    username = e1.get()
    global password
    password = e2.get()
    try:
        jira = JIRA(
            basic_auth = (username, password),
            #options = {'server': 'http://142.104.193.65:8080'}
            options = {'server': 'https://jira.oceannetworks.ca/'}
        )
        log_inWindow.destroy()
    except:
        mb.showerror("Error", "Login Unsuccessfull")
    
tk.Button(log_inWindow, text = 'Login', font = "bold", command = save_textvariable, bd = 4, width = 8, bg = "#0E69C1", fg = "white", activeforeground = "gray").grid(row = 2, column = 2, padx = 10,)

log_inWindow.mainloop()

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~Second Window
"""Creates main GUI and asks for user to browse and select excel workbook to be used
 
    Lets user choose excel file from file explorer and only accepts openpyxl file formats"""

initWindow = tk.Tk()

initWindow.title('Open Excel File')
initWindow.geometry("500x300")
initWindow.resizable(False, False)
initWindow['bg'] = "#0E69C1"

labelframe1 = LabelFrame(initWindow, text="Choose Excel Workbook", bg = "#0E69C1", fg = "white")
labelframe1.pack(fill= "both", expand="yes", padx = 10, pady = 10) 

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def OpenFile():
    """ Open File explorer and lets user select exsisting excel workbook and worksheet to be used """
    global mypath
    mypath = filedialog.askopenfilename(initialdir = "C:",
                           filetypes = (("Excel Workbook", "*.xlsx"), ("Excel Macro-Enabled Workbook","*.xlsm")),
                           title = "Choose a file."
                           )

    global workbookTitle
    workbookTitle = os.path.basename(mypath)
    wbkTitle = StringVar()
    wbkTitle.set(workbookTitle)
    

    Entry(labelframe1, textvariable = wbkTitle, state = DISABLED, width = 35, font = 'bold').place(x = 20, y = 25)


def quit_initWindow():
    initWindow.destroy()

def create_ticket(row):
    # Assign the values
    __instrument = row['Instrument']
    __serialNumber = row['Serial Number']
    __deviceID = row['Device ID']

    __location = row['Node']
    __communications = row['Communications']
    __cableID = row['Cable']
    __instrumentCategory = row['Instrument Category']
    __ipAddress = row['IP Address']

    __components = row['Component']
    __junctionBox = row['Junction Box']
    __port = row['Port']
    __outwardIssue = row['Linked To']

    if(row['Operation'] == 'Deploy'):
        __summaryTitle = 'Instrument Qualification'
    if(row['Operation'] == 'Recover'):
        __summaryTitle = 'Instrument Recovery'

    # Connect to jira
    # Authentication done by using username and password
    #username = 'mtcelec2'
    #password = '1q2w3e4R!'

    jira = JIRA(
        basic_auth = (username, password),
        #options = {'server': 'http://142.104.193.65:8080'}
        options = {'server': 'https://jira.oceannetworks.ca/'}
    )

    #create a ticket for current row
    new_issue = jira.create_issue(
        project = {'key': 'EN'}, 
        summary = "%s: %s SI: %s DI: %s" % (__summaryTitle, __instrument, __serialNumber, __deviceID),
        description = "Site Location: %s\n Communications: %s\n Cable EXT ID: %s\n Instrument Category: %s\n IP Address: %s\n Junction Box: %s\n Port: %s\n" % (__location, __communications, __cableID, __instrumentCategory, __ipAddress, __junctionBox, __port), 
        issuetype = {'name': 'Task'}, 
        components = [{'name' : __components}],
        customfield_10794 = {'id': "10453"},            # Bill of work to Customers (Default: ONC Internal)
        customfield_10592 = "%s" % __serialNumber,      # Serial # field
        customfield_10070 = __deviceID)                 # Device ID field

    # add the linkedto
    
    if(isinstance(row['Linked To'], str)):
        jira.create_issue_link("Related", new_issue.key, __outwardIssue, None)

    return new_issue.key

def update_ticket(row):
    # Assign the values
    __instrument = row['Instrument']
    __serialNumber = row['Serial Number']
    __deviceID = row['Device ID']

    __location = row['Node']
    __communications = row['Communications']
    __cableID = row['Cable']
    __instrumentCategory = row['Instrument Category']
    __ipAddress = row['IP Address']

    __components = row['Component']
    __junctionBox = row['Junction Box']
    __port = row['Port']
    __outwardIssue = row['Linked To']
    __summaryTitle = 'Instrument Qualification'

    __issueLink = row['Work Ticket']

    findEN = 'EN'
    __IssueKey = __issueLink[__issueLink.find(findEN):]

    # connect to Jira
    jira = JIRA(
        basic_auth = (username, password),
        #options = {'server': 'http://142.104.193.65:8080'}
        options = {'server': 'https://jira.oceannetworks.ca/'}
    )

    # Issue to be updated
    issue = jira.issue(__IssueKey)

    # Update the issue
    issue.update(
        project = {'key': 'EN'}, 
        summary = "%s: %s SI: %s DI: %s" % (__summaryTitle, __instrument, __serialNumber, __deviceID),
        description = "Site Location: %s\n Communications: %s\n Cable EXT ID: %s\n Instrument Category: %s\n IP Address: %s\n Junction Box: %s\n Port: %s\n" % (__location, __communications, __cableID, __instrumentCategory, __ipAddress, __junctionBox, __port), 
        customfield_10794 = {'id': "10453"},            # Bill of work to Customers (Default: ONC Internal)
        customfield_10592 = "%s" % __serialNumber,      # Serial # field
        customfield_10070 = __deviceID)                 # Device ID field


def ExtractRow():
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Here is extraction part
    # which extract the deploy rows from the oringial worksheets
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    global df_out

    # import the excel file into a dataframe
    df_whole = pd.read_excel(workbookTitle, sheet_name = None)

    #print(df_whole)

    # count how many sheets in the whole excel sheet
    sheetCnt = len(df_whole.keys())

    # create a list and initialize the list item in the list
    # for storing the subsheets
    df_sheet = []
    for i in range(0,sheetCnt):
        df_sheet.append("df_{}".format(i))

    # create a list and append in the node name for filling use
    node_fill = []
    for sheet in df_whole.keys():
        node_fill.append(sheet)

    # open the sheet seperately and store them seperately in the created list
    for i in range(0,sheetCnt):
        df_sheet[i] = pd.read_excel(workbookTitle, sheet_name = i)

    global site_dict
    for site in node_fill:    
        site_dict[site] = ''

    # set a variable for storing the current position
    pos = 0

    # loop through all sheets in the excel sheets
    for sheet in df_sheet:
        # add a column for storing the row number for tracing back
        sheet.insert(14, "RowNum", np.nan)

        # Drop unused columns
        # only keep column from column 1 to column 14
        # which is NODE to RowNum
        sheet.drop(sheet.iloc[:, 14::], inplace = True, axis = 1)
        
        # grab the site EN number and store to dictionary
        site_dict[node_fill[pos]] = sheet.columns[0]
        
        # Assign the column names
        # since the first row initially is empty
        sheet.columns = ['Node', 'Junction Box', 'Port','Communications','Cable to Connector Panel', 'Cable', 'IP Address',
                    'Instrument Category', 'Instrument', 'Serial Number', 'Device ID', 'Work Ticket', 'Operation','RowNum']
    
        # Drop the first row
        sheet = sheet.drop(index=0)
    
        # ~~~~~~~~~~~~~~~~~~~deal with the sub ticket
        # Set up the intial fill-in value
        currPort = ''
        currCom = ''
        currCtCP = ''
        currCEI = ''
        currIP = ''
        currIC = ''
        currIns = ''
        currSN = ''
        currDI = ''
        currWT = ''
        currOp = ''
    
        # check in the port column, whether already see the port in string already
        # initially set to false
        meetString = False
    
        # iterate the rows in the subsheet
        # and record the values when encouter to the first valid record
        for index, row in sheet.iterrows():
            if(isinstance(row['Port'], str)):
                currPort = row['Port']
                currCom = row['Communications']
                currCtCP = row['Cable to Connector Panel']
                currCEI = row['Cable']
                currIP = row['IP Address']
                currIC = row['Instrument Category']
                currIns = row['Instrument']
                currSN = row['Serial Number']
                currDI = row['Device ID']
                currWT = row['Work Ticket']
                currOp = row['Operation']
                break
            
        # iterate rhw rows in the subsheet and do the subtickts fill        
        for index, row in sheet.iterrows():
            # if meet a empty port cell and already seen a string port before
            # check whether this row is empty or not
            # if is not, fill in with current fill-in values
            
            # fill in the row number
            row['RowNum'] = index+2

            if(meetString == True and (not isinstance(row['Port'], str))):
            
                isCom = isinstance(row['Communications'], str)
                isCtCP = isinstance(row['Cable to Connector Panel'], str)
                isCEI = isinstance(row['Cable'], str)
                isIP = isinstance(row['IP Address'], str)
                isIC = isinstance(row['Instrument Category'], str)
                isIns = isinstance(row['Instrument'], str)
                isSN = isinstance(row['Serial Number'], str)
                isDI = isinstance(row['Device ID'], str)
                isWT = isinstance(row['Work Ticket'], str)
                isOp = isinstance(row['Operation'], str)
            
                # check whether the current is empty or not
                # in other words, the possibility of being a subtickets
                isNotEmpty = isCom or isCtCP or isCEI or isIP or isIC or isIns or isSN or isDI or isWT or isOp
               
    #        isNotEmpty = row['Communications'] != np.nan or row['Cable to Connector Panel'] != np.nan 
    #        or row['Cable EXT ID'] != np.nan or row['IP Address'] != np.nan or row['IP Address'] != np.nan 
    #        or row['Instrument Category'] != np.nan or row['Instrument'] != np.nan or row['Serial Number'] != np.nan 
    #        or row['Device ID'] != np.nan or row['Work Ticket'] != np.nan or row['Operation'] != np.nan

                # if is not empty -> is a subtickets -> fill the empty cells using current fill-in values
                # note: the fill-in values is inherited from the last port-non-empty row
                if(isNotEmpty):
            
                    # fill out the empty cells of children tickets
                    row['Port'] = currPort
                        
                    if(not isCom):
                        row['Communications'] = currCom
                
                    #if row['Communications'] == np.nan:
                    #   row['Communications'] = currCom
                
                    if (not isCtCP):
                        row['Cable to Connector Panel'] = currCtCP
                
                    if (not isCEI):
                        row['Cable'] = currCEI
                
                    if (not isIP):
                        row['IP Address'] = currIP
                
                    if (not isIC):
                        row['Instrument Category'] = currIC
                
                    if (not isIns):
                        row['Instrument'] = currIns
                
                    if (not isSN):
                        row['Serial Number'] = currSN
                
                    if (not isDI):
                        row['Device ID'] = currDI
                    
                    # do not inherit the work ticket value
                    #if (not isWT):
                        #row['Work Ticket'] = currWT
                
                    if (not isOp):
                        row['Operation'] = currOp
                    
            # if the current row is a new ticket, assign the current row values to the fill-in values
            if(isinstance(row['Port'], str) and row['Port'] != currPort):
                meetString = True
                currPort = row['Port']
                currCom = row['Communications']
                currCtCP = row['Cable to Connector Panel']
                currCEI = row['Cable']
                currIP = row['IP Address']
                currIC = row['Instrument Category']
                currIns = row['Instrument']
                currSN = row['Serial Number']
                currDI = row['Device ID']
                currWT = row['Work Ticket']
                currOp = row['Operation']
                   
        junction_box_fill = ""
        # find the juction box fill
        for item in sheet['Junction Box']:
            thisType = isinstance(item, str) 
            if thisType:
                junction_box_fill = item
        
        # fill the Node column and Junction Box column            
        sheet['Node'] = sheet['Node'].fillna(node_fill[pos])
        sheet['Junction Box'] = sheet['Junction Box'].fillna(junction_box_fill)
    
        # Extract the deploy row and update row from the current sheet and append the row into the output dataframe
        for index, row in sheet.iterrows():
            if(row["Operation"] == "Deploy" and (not isinstance(row['Work Ticket'], str))):
                df_out = df_out.append(row, ignore_index = True)[list(sheet)]

            if(row["Operation"] == "Update" and isinstance(row['Work Ticket'], str)):
                df_out = df_out.append(row, ignore_index = True)[list(sheet)]

            if(row["Operation"] == "Recover" and (not isinstance(row['Work Ticket'], str))):
                df_out = df_out.append(row, ignore_index = True)[list(sheet)]

        #jump to the next sheet
        pos += 1

    try:
        # insert the column needed in the jira import tool
        df_out.insert(11, "Component", "Test and Development")
        df_out.insert(13, "Linked To", np.nan)
        quit_initWindow()
    except:
        # no records need to generate tickets
        mb.showerror("Error", "Nothing to generate.")
        return
    """
    # using dictionary to convert specific columns 
    # not feasibile, becasue some cells are empty
    convert_dict = {'Serial Number': int, 
                    'Device ID': int} 
  
    df_out = df_out.astype(convert_dict)
    """
    df_out['Device ID'] = df_out['Device ID'].astype(int)

    # iterate throght the dataframe and change the data type
    for index, row in df_out.iterrows():
        """
        # check whether SN only contain number
        if(row['Serial Number'].isdigit()):
            row['Serial Number'] = int(row['Serial Number'])
        """
        #print(df_out['Device ID'][index])
        # change the device ID to int
        # This is Wrong again!!! row['Device ID'] = int(row['Device ID'])
        # df_out['Device ID'][index] = int(df_out['Device ID'][index]) -> will give me error

        # disable chained assignments
        pd.options.mode.chained_assignment = None
        
        if(isinstance(row['Serial Number'], float)):
            try:
                df_out['Serial Number'][index] = int(df_out['Serial Number'][index])
            except:
                continue

        #print(type(df_out['Serial Number'][index]))
    
    #print(df_out)

 
def AutomationRow():
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Here is automation part
    # which generate the tickets automatically
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    # drop the selected row that don't want to create ticket
    global df_out
    for ctr, int_var in enumerate(cb_intvar):
        if int_var.get():
            df_out = df_out.drop(index = ctr)

    # suppress the warning
    pd.set_option('mode.chained_assignment', None)

    # initialzie the variable
    __instrument = ''
    __serialNumber = ''
    __deviceID = ''

    # Descriptions
    __location = ''
    __communications = ''
    __cableID = ''
    __instrumentCategory = ''
    __ipAddress = ''
    __junctionBox = ''
    __port = ''

    __components = ''
    __outwardIssue = ''
    __summaryTitle = ''

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    # initialize the current parent EN number
    currEN = ''

    # assign the first value of the port column to currEN
    # for comparing if it is a parent or a child
    currPort = df_out.iloc[0]['Port']
    meetInitialPort = False

    for index, row in df_out.iterrows():
        if(row['Operation'] == 'Update'):
            # update the ticket
            update_ticket(row)

        elif(row['Operation'] == 'Deploy' or row['Operation'] == 'Recover'):
            # create the ticket
            if(row['Port'] == currPort and meetInitialPort == True):
                # This is a child row
                # check whether has the ticket already
                # if yes, update the ticket with linkedto, remain the currEN and currPort the same 
                # if no, fill the linkedto(both row and df) with currEN and create a new ticket and write into dataframe
        
                # fill the linkedTo with currEN
                df_out['Linked To'][index] = currEN
                row['Linked To'] = currEN
                    
                # create a new jira ticket
                myKey = create_ticket(row)
                    
                # write the new ticket URL into dataframe
                df_out['Work Ticket'][index] = "https://jira.oceannetworks.ca/browse/%s" % myKey

                # df_out['Work Ticket'][index] = "http://142.104.193.65:8080/browse/%s" % myKey
                # This IS WRONG: row['Work Ticket'] = "http://142.104.193.65:8080/browse/%s" % myKey. because it won't change the df itself
                    
                # do not need to update the EN number or currPort because this is a child ticket
                    
                
            # This is row one
            if(row['Port'] == currPort and meetInitialPort == False):
                # Mark as already seen the first row
                meetInitialPort = True
                
                # link to the site, create a new ticket, write into the dataframe and record the EN number

                # loop through the dictionary, and write the site EN number to both row and df
                for mysite, siteEN in site_dict.items(): 
                    if (mysite == row['Node']):
                        df_out['Linked To'][index] = siteEN
                        row['Linked To'] = siteEN
                        
                myKey = create_ticket(row)
                    
                # write the new ticket URL into dataframe
                df_out['Work Ticket'][index] = "https://jira.oceannetworks.ca/browse/%s" % myKey
            
                # THIS IS WRONG: row['Work Ticket'] = "http://142.104.193.65:8080/browse/%s" % myKey, same as the last one
                    
                # record EN number
                currEN = myKey

                # no need to record the current port since this is the first and I already record it

            if(row['Port'] != currPort):
                # this means this row is a new parent row
                # link to the site, create a new ticket, write the URL to the dataframe, update the currEN, update the currPort
                
                # loop through the dictionary, and write the site EN number to both row and df
                for mysite, siteEN in site_dict.items(): 
                    if (mysite == row['Node']):
                        df_out['Linked To'][index] = siteEN
                        row['Linked To'] = siteEN
                        
                # create a ticket, write the URL to dataframe and update the currEN, update the currPort
                myKey = create_ticket(row)
                    
                # write the new ticket URL into dataframe
                df_out['Work Ticket'][index] = "https://jira.oceannetworks.ca/browse/%s" % myKey
            
                # update EN number
                currEN = myKey

                # record the current port
                currPort = row['Port']

    #print(df_out)
 
    # iterarte through the sheet and fill back to oringinal cruise sheet
    for index, row in df_out.iterrows():
        # load cruise configuration sheet
        wb = load_workbook(mypath)

        sheetName = row['Node']
        rowNumber = row['RowNum']

        ws = wb[sheetName]
        
        if (row['Operation'] == 'Deploy'):
            # grab the cell
            wcell_Ticket = ws.cell(rowNumber, 12)
            wcell_Operation = ws.cell(rowNumber, 13)

            # change the cell
            wcell_Ticket.value = row['Work Ticket']
            wcell_Operation.value = 'Deploy(created)'

        elif(row['Operation'] == 'Update'):
            # grab the cell
            wcell_Operation = ws.cell(rowNumber, 13)
            wcell_Operation.value = 'Updated!'

        elif(row['Operation'] == 'Recover'):
            # grab the cell
            wcell_Ticket = ws.cell(rowNumber, 12)
            wcell_Operation = ws.cell(rowNumber, 13)

            # change the cell
            wcell_Ticket.value = row['Work Ticket']
            wcell_Operation.value = 'Recover(created)'

        # save the file
        wb.save(mypath)

    try:
        df_out.to_excel(r'Deploy.xlsx', index = False)
        mb.showinfo('Issue(s) Created/Updated', 'Deploy.xlsx successfully generated.' )
    except:
        mb.showerror("Error", "Creation Unsuccessfull")

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~Second Window
tk.Button(labelframe1, text = "BROWSE", command = OpenFile, font = 'bold', bd = 4, width = 10, bg = "white", fg = "black", activeforeground = "gray").place(x = 350, y = 25)
tk.Button(labelframe1, text = "ENTER", command = ExtractRow, font = 'bold', bd = 4, width = 10, bg = "white", fg = "black", activeforeground = "gray").place(x = 350, y = 100)

initWindow.mainloop()
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""
def mytest():
    global df_out
    for ctr, int_var in enumerate(cb_intvar):
        if int_var.get():
            df_out = df_out.drop(index = ctr)

    print(df_out)  
"""    
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~Third Window
# Now give a preview of the df_out to let user to change!!!
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

"""Creates the main window where the selecting, creating, and updating issues are done"""

mainWindow = tk.Tk()
mainWindow.title("Import Row")
mainWindow.state('zoomed')
mainWindow.configure(bg = "#0E69C1")


mainWindow.columnconfigure(0, weight = 1)
mainWindow.rowconfigure(0, weight = 1)

canvas = tk.Canvas(mainWindow) 
frame = tk.Frame(canvas)

total_column = df_out.shape[1]
total_row = df_out.shape[0]

create_button = tk.Button(frame, text = "CREATE",  command = AutomationRow, bd = 4, relief = RAISED, width = 15, height = 2)

create_button.grid(row = 0, column = 1, padx = 15, pady = 10)

tk.Label(frame, text = "Check the row NOT to create ticket.", font = 'bold').grid(row = 0, column = 3, padx = 10, pady = 10)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# show df_out on GUI
# print(df_out)

start = 1
for col in df_out.columns: 
    tk.Label(frame, text = col, font = 'bold').grid(row = 1, column = start, padx = 10, pady = 10)
    start += 1

for j in range(0, total_row):  
    for i in range(0, total_column):
        dataCell = df_out.iloc[j,i]
        tk.Label(frame, text = dataCell).grid(row = j+2, column = i+1, padx = 10, pady = 10)

 # Checkboxes created to select row of data to grab data from        
cb_intvar = []  
for j in range(1, total_row+1):  
    cb_intvar.append(IntVar())
    chbx = tk.Checkbutton(frame, variable=cb_intvar[-1])
    chbx.grid(row = j+1, column = 0, sticky = 'w')

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# """Block of code for the scrollbar"""

def on_resize(event):
    """Resize canvas scrollregion when the canvas is resized."""
    canvas.configure(scrollregion=canvas.bbox('all'))

canvas.create_window(0, 0, anchor = 'nw', window = frame)

vbar = ttk.Scrollbar(mainWindow, orient = 'vertical', command = canvas.yview)
hbar = ttk.Scrollbar(mainWindow, orient = 'horizontal', command = canvas.xview)
canvas.configure(xscrollcommand = hbar.set,
                yscrollcommand = vbar.set,
                scrollregion = canvas.bbox('all'))

canvas.grid(row = 0, column = 0, sticky = 'eswn')
vbar.grid(row = 0, column = 1, sticky = 'ns')
hbar.grid(row = 1, column = 0, sticky = 'ew')

canvas.bind('<Configure>', on_resize)

mainWindow.mainloop()

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~